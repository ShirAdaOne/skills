#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
导出LP语音控车测试用例到Excel
"""
import re
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def parse_markdown_table(md_content):
    """解析Markdown中的所有表格，提取测试用例数据"""
    all_rows = []
    lines = md_content.split('\n')
    
    in_table = False
    header_done = False
    
    for i, line in enumerate(lines):
        line = line.strip()
        
        if not line.startswith('|'):
            in_table = False
            header_done = False
            continue
        
        # 跳过分隔行
        if re.match(r'^\|[\s\-\|:]+\|$', line):
            if in_table:
                header_done = True
            continue
        
        cells = [c.strip() for c in line.split('|')]
        cells = [c for c in cells if c != '']  # 去除首尾空元素
        
        if not cells:
            continue
        
        # 检查是否是表头行（第一列包含"用例编号"）
        if '用例编号' in cells[0]:
            in_table = True
            header_done = False
            continue
        
        if in_table and header_done and cells:
            # 检查第一列是否是用例编号格式
            if cells[0].startswith('BFO-HMI-'):
                # 替换<br>为换行符
                cells = [c.replace('<br>', '\n') for c in cells]
                all_rows.append(cells)
    
    return all_rows

def export_to_excel(md_file, excel_file, sheet_name):
    """将Markdown测试用例导出到Excel"""
    
    # 读取MD文件
    with open(md_file, 'r', encoding='utf-8') as f:
        md_content = f.read()
    
    # 解析所有用例
    rows = parse_markdown_table(md_content)
    print(f"解析到 {len(rows)} 条用例")
    
    # 创建Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # 表头
    headers = [
        '用例编号', '功能类型', '分组', '用例分级', '用例名称',
        '预置条件', '预置条件-信号描述', '测试步骤', '测试步骤-信号描述',
        '预期结果', '预期结果-信号描述', '标签信息', '备注', '层级'
    ]
    
    # 设置表头样式
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=10)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    ws.row_dimensions[1].height = 30
    
    # P0/P1 填充颜色
    p0_fill = PatternFill(start_color='E8F4FD', end_color='E8F4FD', fill_type='solid')
    p1_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
    
    normal_font = Font(name='微软雅黑', size=9)
    normal_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    center_alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    
    # 写入数据
    for row_idx, row_data in enumerate(rows, 2):
        # 确保14列
        while len(row_data) < 14:
            row_data.append('-')
        row_data = row_data[:14]
        
        priority = row_data[3] if len(row_data) > 3 else 'P0'
        row_fill = p0_fill if priority == 'P0' else p1_fill
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = normal_font
            cell.border = thin_border
            cell.fill = row_fill
            
            # 居中列：编号、功能类型、分组、用例分级、标签、层级
            if col_idx in [1, 2, 3, 4, 12, 14]:
                cell.alignment = center_alignment
            else:
                cell.alignment = normal_alignment
        
        # 设置行高（根据内容多少）
        max_lines = max([len(str(v).split('\n')) for v in row_data] + [1])
        ws.row_dimensions[row_idx].height = max(25, min(max_lines * 16, 120))
    
    # 设置列宽
    col_widths = [20, 10, 10, 10, 28, 30, 28, 30, 28, 38, 38, 14, 20, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    # 保存
    wb.save(excel_file)
    print(f"Excel文件已保存：{excel_file}")
    return len(rows)

if __name__ == '__main__':
    import os
    os.chdir('/home/dr/codetree/hmi-testcase-auto-generation')
    
    md_file = 'LP_VoiceControl_TestCases.md'
    excel_file = 'LP_VoiceControl_TestCases.xlsx'
    sheet_name = 'LP语音控车测试用例'
    
    print('='*60)
    print('LP语音控车测试用例 - Excel导出工具')
    print('='*60)
    
    count = export_to_excel(md_file, excel_file, sheet_name)
    
    print(f'\n✅ 导出完成！共 {count} 条用例')
    print(f'输出文件：{excel_file}')
    
    # 统计P0/P1
    import re
    with open(md_file, 'r', encoding='utf-8') as f:
        content = f.read()
    
    all_ids = re.findall(r'BFO-HMI-LP-VC-\d+', content)
    unique_ids = set(all_ids)
    p0_count = len(re.findall(r'\| P0 \|', content))
    p1_count = len(re.findall(r'\| P1 \|', content))
    
    print(f'\n📊 统计信息：')
    print(f'  - MD用例总数：{len(unique_ids)} 个')
    print(f'  - Excel导出：{count} 条')
    print(f'  - P0用例：{p0_count} 个')
    print(f'  - P1用例：{p1_count} 个')
    
    if count == len(unique_ids):
        print(f'\n✅ 数量一致，导出成功！')
    else:
        print(f'\n⚠️  数量差异：MD={len(unique_ids)}，Excel={count}，请检查MD格式')
